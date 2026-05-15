import streamlit as st
import pandas as pd
import io
import re
import os
import datetime
from openpyxl import load_workbook

# ==========================================
# 💎 UI: ULTRA-PREMIUM GLASSMORPHISM (ยึดตามธีมเดิมที่คุณส่งมา)
# ==========================================
st.set_page_config(page_title="QAD System Hub", layout="wide")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Kanit:wght@300;500;700&family=Inter:wght@400;600;900&display=swap');

    /* พื้นหลัง Animated Gradient ทรงพลัง (ยึดตามธีมเดิม) */
    .stApp {
        background: linear-gradient(-45deg, #f8fafc, #e2e8f0, #f1f5f9, #cbd5e1);
        background-size: 400% 400%;
        animation: gradient 15s ease infinite;
        color: #0f172a;
        font-family: 'Inter', 'Kanit', sans-serif;
    }

    @keyframes gradient {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }

    /* Sidebar แบบกระจกฝ้าโปร่งแสง */
    [data-testid="stSidebar"] {
        background-color: rgba(255, 255, 255, 0.6) !important;
        backdrop-filter: blur(15px);
        border-right: 1px solid rgba(255, 255, 255, 0.3);
    }

    /* ปุ่มเมนูหลัก: หรูหรา มีมิติ (Glassmorphism) */
    div.stButton > button:first-child {
        background: rgba(255, 255, 255, 0.7);
        color: #0f172a;
        border: 1px solid rgba(15, 23, 42, 0.1);
        border-radius: 0px; 
        height: 140px;
        font-size: 30px !important;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 1px;
        transition: all 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275);
        box-shadow: 0 10px 30px rgba(0,0,0,0.05), inset 0 0 0 1px rgba(255,255,255,0.8);
        backdrop-filter: blur(10px);
    }

    div.stButton > button:hover {
        background: #0f172a;
        color: #ffffff !important;
        transform: translateY(-5px) scale(1.01);
        box-shadow: 0 20px 40px rgba(15, 23, 42, 0.2);
        border: 1px solid #3b82f6;
    }

    /* ชื่อแอปหลัก */
    .main-title {
        font-family: 'Kanit', sans-serif;
        font-size: 80px;
        font-weight: 700;
        color: #0f172a;
        text-align: center;
        line-height: 1.1;
        margin-bottom: 10px;
    }

    .sub-title {
        text-align: center;
        color: #475569;
        font-size: 24px;
        font-weight: 500;
        margin-bottom: 3rem;
    }

    /* ขนาดตัวหนังสือคงเดิม */
    p, label, .stMarkdown {
        font-size: 19px !important;
    }
    
    h3 {
        font-size: 28px !important;
        font-weight: 700 !important;
        color: #1e293b !important;
    }

    /* Tabs สไตล์ Glass */
    .stTabs [data-baseweb="tab"] {
        height: 60px;
        font-size: 20px;
        background-color: rgba(226, 232, 240, 0.5);
        border-radius: 0px;
        color: #475569;
        padding: 0 40px;
        backdrop-filter: blur(5px);
    }

    .stTabs [aria-selected="true"] {
        background-color: #0f172a !important;
        color: #ffffff !important;
    }

    /* Table & Data Handling */
    [data-testid="stDataFrame"] {
        border: 2px solid #0f172a !important;
        background: white;
    }
    
    /* File Uploader แบบ Industrial Glass */
    .stFileUploader section {
        border: 2px dashed #0f172a;
        border-radius: 0px;
        background-color: rgba(255, 255, 255, 0.5);
    }
    </style>
""", unsafe_allow_html=True)

if "current_app" not in st.session_state:
    st.session_state.current_app = "Main Menu"

def go_to_menu():
    for key in list(st.session_state.keys()):
        if key not in ["current_app", "reset_counter", "uploader_key"]:
            del st.session_state[key]
    st.session_state.current_app = "Main Menu"
    st.rerun()

# ==========================================
# APP 1: FILE VALIDATOR (ยึดตามโค้ดเดิมของคุณ 100%)
# ==========================================
def app_file_validator():
    st.markdown("<h1 style='text-align: center; color: #0f172a; font-size: 50px;'>📁 File Validator</h1>", unsafe_allow_html=True)
    st.markdown("<p class='center-text' style='text-align: center; font-size: 22px; color: #64748b;'>ตรวจสอบ format ไฟล์ก่อนส่ง ✨</p>", unsafe_allow_html=True)
    st.divider()

    with st.sidebar:
        st.markdown("### 🧭 Control Panel")
        if st.button("🏠 Home Menu", use_container_width=True):
            go_to_menu()
        if st.button("🔄 Refresh System", use_container_width=True):
            for key in list(st.session_state.keys()):
                if key != "current_app": del st.session_state[key]
            st.session_state.reset_counter = st.session_state.get('reset_counter', 0) + 1
            st.rerun()

    if 'reset_counter' not in st.session_state:
        st.session_state.reset_counter = 0

    TARGET_SHEET = "RAMP v1.3"

    def get_column_letter(n):
        result = ""
        while n >= 0:
            result = chr(n % 26 + 65) + result
            n = n // 26 - 1
        return result

    def get_available_models():
        files = [f for f in os.listdir('.') if f.startswith('reference_') and (f.endswith('.xlsx') or f.endswith('.xlsm'))]
        return {f.replace('reference_', '').split('.')[0]: f for f in files}

    try:
        available_models = get_available_models()
        model_list = ["-- เลือกโมเดลอ้างอิง --"] + sorted(list(available_models.keys()))
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**1. เลือกโมเดลอ้างอิง**")
            selected_model_name = st.selectbox("Model", model_list, index=0, key=f"v_sel_{st.session_state.reset_counter}", label_visibility="collapsed")
        
        if selected_model_name != "-- เลือกโมเดลอ้างอิง --":
            with c2:
                st.markdown("**2. อัปโหลดไฟล์ที่ต้องการตรวจ**")
                uploaded_file = st.file_uploader("Upload", type=["xlsx", "xlsm"], key=f"v_up_{st.session_state.reset_counter}", label_visibility="collapsed")

            if uploaded_file:
                st.write("---")
                with st.status("🔍 กำลังตรวจสอบความถูกต้อง...", expanded=True) as status:
                    wb = load_workbook(uploaded_file, data_only=False)
                    ws = wb[TARGET_SHEET]
                    df_ref = pd.read_excel(available_models[selected_model_name], sheet_name=TARGET_SHEET, header=None).fillna("")
                    df_user = pd.read_excel(uploaded_file, sheet_name=TARGET_SHEET, header=None).fillna("")
                    
                    f_errors, missing_data, extra_data, d_errors, k_errors = [], {}, {}, [], []

                    for r, c, label in [(2, 5, "F3"), (4, 5, "F5")]:
                        if str(df_user.iloc[r, c]).strip() != str(df_ref.iloc[r, c]).strip():
                            f_errors.append({"Position": label, "Found": df_user.iloc[r, c], "Target": df_ref.iloc[r, c]})

                    for r in range(76):
                        for c in range(df_ref.shape[1]):
                            if r >= 12 and c in [3, 10]: continue
                            ref_v, user_v = str(df_ref.iloc[r, c]).strip(), str(df_user.iloc[r, c]).strip()
                            if ref_v != "" and (user_v == "" or user_v == "nan"):
                                missing_data.setdefault(get_column_letter(c), []).append(str(r+1))
                            elif ref_v == "" and (user_v != "" and user_v != "nan") and r+1 != 12:
                                extra_data.setdefault(get_column_letter(c), []).append(str(r+1))

                    for row_idx in range(13, 77): 
                        for col_idx, (col_label, error_list) in enumerate(zip(['D', 'K'], [d_errors, k_errors])):
                            target_col = 4 if col_label == 'D' else 11
                            cell = ws.cell(row=row_idx, column=target_col)
                            if cell.value is not None:
                                val, fmt = cell.value, str(cell.number_format).lower()
                                has_format = (('y' in fmt or ('d' in fmt and 'm' in fmt)) and 'h' in fmt)
                                is_valid_data = False
                                if isinstance(val, (datetime.datetime, datetime.date)):
                                    is_valid_data = True
                                else:
                                    try:
                                        dt_temp = pd.to_datetime(val)
                                        is_valid_data = False if (dt_temp.hour == 0 and dt_temp.minute == 0 and dt_temp.second == 0) else True
                                    except: is_valid_data = False
                                if not has_format or not is_valid_data:
                                    reason = "❌ Format ผิด" if not has_format else "❌ ขาดเวลา"
                                    error_list.append({"Row": row_idx, "Value": str(val), "Status": reason})
                    status.update(label="ตรวจสอบเสร็จเรียบร้อย!", state="complete", expanded=False)

                if not (f_errors or missing_data or extra_data or d_errors or k_errors):
                    st.balloons()
                    st.success("✨ ยินดีด้วย! ข้อมูลถูกต้องสมบูรณ์ 100%")
                else:
                    if f_errors:
                        st.error("⚠️ part no. / drawing no. ไม่ถูกต้อง (ช่อง F3/F5)")
                        st.table(pd.DataFrame(f_errors))
                    col_mis, col_ext = st.columns(2)
                    with col_mis:
                        if missing_data:
                            st.warning("📉 ข้อมูลที่หายไป (Missing Data)")
                            st.table([{"Column": k, "Rows": ", ".join(v)} for k, v in missing_data.items()])
                    with col_ext:
                        if extra_data:
                            st.warning("🚫 ข้อมูลส่วนเกิน (Extra Data)")
                            st.table([{"Column": k, "Rows": ", ".join(v)} for k, v in extra_data.items()])
    except Exception as e: st.error(f"เกิดข้อผิดพลาด: {e}")

# ==========================================
# APP 2: WASHING DATE PROCESSOR (แก้ไข Logic ให้ทำงานได้จริงตาม Logic ดั้งเดิม)
# ==========================================
def read_excel_washing(file):
    try: return pd.read_excel(file, engine="openpyxl", header=None)
    except: return pd.read_excel(file, engine="xlrd", header=None)

def read_file1_washing(file):
    df = read_excel_washing(file)
    col, start_row = 5, 16 # คอลัมน์ F แถว 17
    data = df.iloc[start_row:, col]
    lot_list = []
    for val in data:
        if pd.isna(val) or str(val).strip() == "": break
        lot_list.append(str(val).strip())
    return pd.DataFrame({"Lot": lot_list})

def read_file2_washing(file):
    df = read_excel_washing(file)
    header_row = None
    for i in range(20):
        row = df.iloc[i].astype(str).str.lower()
        if row.str.contains("runcard").any() and row.str.contains("barcode").any():
            header_row = i; break
    if header_row is None:
        st.error("❌ หา header ไม่เจอ (Runcard / Barcode)"); return pd.DataFrame()
    df.columns = df.iloc[header_row]
    df = df[header_row + 1:]
    df.columns = df.columns.astype(str).str.strip().str.lower()
    lot_cols = [c for c in df.columns if "runcard" in str(c).lower()]
    barcode_cols = [c for c in df.columns if "barcode" in str(c).lower()]
    packed_col = next((c for c in df.columns if "packed" in str(c).lower() and "date" in str(c).lower()), None)
    
    if not lot_cols or not barcode_cols or packed_col is None:
        st.error("❌ หา column ไม่เจอ"); return pd.DataFrame()
    
    df_out = df[[lot_cols[0], barcode_cols[0], packed_col]].copy()
    df_out.columns = ["Lot", "Barcode No", "Packed Date"]
    df_out = df_out.dropna(subset=["Lot"])
    df_out["Lot"] = df_out["Lot"].astype(str).str.strip()
    df_out["Packed Date"] = pd.to_datetime(df_out["Packed Date"], errors="coerce")
    return df_out

def extract_ww_day(barcode):
    try:
        s = str(barcode)
        match = re.search('[A-Za-z]', s)
        if not match: return None, None
        start = match.start()
        code = s[start+3:start+6]
        if len(code) != 3 or not code.isdigit(): return None, None
        return int(code[:2]), int(code[2])
    except: return None, None

def load_database():
    df = pd.read_csv("database.txt")
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%b-%Y", errors="coerce")
    return df

def find_best_date(row, date_db):
    if pd.isna(row["WW"]) or pd.isna(row["Day"]) or pd.isna(row["Packed Date"]): return None
    candidates = date_db[(date_db["WW"] == row["WW"]) & (date_db["Day"] == row["Day"])].copy()
    if candidates.empty: return None
    candidates["diff"] = (candidates["Date"] - row["Packed Date"]).abs()
    return candidates.sort_values("diff").iloc[0]["Date"]

def app_washing_processor():
    st.markdown("<h1 style='text-align: center; color: #0f172a; font-size: 50px;'>📊 Washing Date Processor</h1>", unsafe_allow_html=True)
    st.markdown("<p class='center-text' style='text-align: center; font-size: 22px; color: #64748b;'>ตรวจสอบวันล้างจาก lot no. ⚡</p>", unsafe_allow_html=True)
    st.divider()
    
    with st.sidebar:
        st.markdown("### 🧭 Control Panel")
        if st.button("🏠 Home Menu", use_container_width=True):
            go_to_menu()
        if st.button("🔄 Refresh System", use_container_width=True):
            st.session_state.output, st.session_state.summary, st.session_state.file = None, None, None
            st.session_state.uploader_key = st.session_state.get('uploader_key', 0) + 1
            st.rerun()

    if "uploader_key" not in st.session_state:
        st.session_state.uploader_key = 0

    col_u1, col_u2 = st.columns(2)
    with col_u1:
        st.markdown("**📂 ไฟล์ที่ 1 (Lot/Serial)**")
        file1 = st.file_uploader("Upload 1", type=["xls", "xlsx", "csv"], key=f"file1_{st.session_state.uploader_key}", label_visibility="collapsed")
    with col_u2:
        st.markdown("**📂 ไฟล์ที่ 2 (Runcard / Barcode)**")
        file2 = st.file_uploader("Upload 2", type=["xls", "xlsx", "csv"], key=f"file2_{st.session_state.uploader_key}", label_visibility="collapsed")

    if st.button("🚀 PROCESS DATA", use_container_width=True):
        if not file1 or not file2:
            st.warning("⚠️ กรุณาอัปโหลดไฟล์ให้ครบ")
        else:
            with st.spinner('กำลังประมวลผล...'):
                try:
                    df1 = read_file1_washing(file1)
                    df2 = read_file2_washing(file2)
                    if not df2.empty:
                        merged = pd.merge(df1, df2, on="Lot", how="left").drop_duplicates(subset=["Lot"])
                        merged[['WW', 'Day']] = merged['Barcode No'].apply(lambda x: pd.Series(extract_ww_day(x)))
                        merged["WW"] = pd.to_numeric(merged["WW"], errors="coerce")
                        merged["Day"] = pd.to_numeric(merged["Day"], errors="coerce")
                        
                        date_db = load_database()
                        date_db["WW"] = pd.to_numeric(date_db["WW"], errors="coerce")
                        date_db["Day"] = pd.to_numeric(date_db["Day"], errors="coerce")
                        
                        merged["Washing Date Raw"] = merged.apply(lambda row: find_best_date(row, date_db), axis=1)
                        output = merged[["Lot", "Barcode No", "WW", "Day", "Washing Date Raw"]].copy()
                        output["Washing Date"] = pd.to_datetime(output["Washing Date Raw"]).dt.strftime("%d-%b-%Y")
                        output = output[output["Lot"].astype(str).str.lower() != "lot/serial"].reset_index(drop=True)
                        
                        summary = output.groupby("Washing Date")["Lot"].count().reset_index().rename(columns={"Lot": "Total Lot"})
                        
                        buffer = io.BytesIO()
                        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                            output.drop(columns=["Washing Date Raw"]).to_excel(writer, index=False, sheet_name="Result")
                            summary.to_excel(writer, index=False, sheet_name="Summary")
                        
                        st.session_state.output = output.drop(columns=["Washing Date Raw"])
                        st.session_state.summary = summary
                        st.session_state.file = buffer.getvalue()
                except Exception as e:
                    st.error(f"เกิดข้อผิดพลาด: {e}")

    if st.session_state.get("output") is not None:
        st.success("✅ Process สำเร็จ!")
        t1, t2 = st.tabs(["📋 Result Detail", "📊 Summary View"])
        with t1: st.dataframe(st.session_state.output, use_container_width=True)
        with t2: st.dataframe(st.session_state.summary, use_container_width=True)
        st.download_button(label="📥 Download Excel", data=st.session_state.file, file_name="washing_date_result.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ==========================================
# MAIN ROUTING
# ==========================================
if st.session_state.current_app == "Main Menu":
    st.markdown("<div style='height: 80px;'></div>", unsafe_allow_html=True)
    st.markdown("<p class='main-title'>QAD SYSTEM HUB</p>", unsafe_allow_html=True)
    st.markdown("<p class='sub-title'>📂 QE Support Application 🚀</p>", unsafe_allow_html=True)
    
    _, col_main, _ = st.columns([0.1, 0.8, 0.1])
    
    with col_main:
        c1, c2 = st.columns(2, gap="large")
        with c1:
            st.markdown("### 📋 ตรวจเช็ค Format ไฟล์")
            if st.button("📁 File Validator", use_container_width=True):
                st.session_state.current_app = "Validator"; st.rerun()
        with c2:
            st.markdown("### 🧼 ตรวจวันล้างสินค้า")
            if st.button("📊 Washing Date Processor", use_container_width=True):
                st.session_state.current_app = "Processor"; st.rerun()

    st.markdown("<div style='height: 100px;'></div>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #94a3b8; font-size: 16px;'>© 2026 Quality Engineering | Systems v2.0 ✨</p>", unsafe_allow_html=True)

elif st.session_state.current_app == "Validator": app_file_validator()
elif st.session_state.current_app == "Processor": app_washing_processor()
